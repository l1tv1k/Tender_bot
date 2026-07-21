import logging
from pathlib import Path

from smart_extractor import SmartDocxExtractor

logger = logging.getLogger(__name__)


class TenderDocumentExtractor:
    """Extracts text from document formats commonly published by EIS."""

    supported_suffixes = {".docx", ".pdf", ".xlsx"}

    def __init__(self):
        self.docx_extractor = SmartDocxExtractor()

    def extract(self, file_path: str) -> str:
        suffix = Path(file_path).suffix.casefold()
        if suffix == ".docx":
            return self.docx_extractor.extract(file_path)
        if suffix == ".pdf":
            return self._extract_pdf(file_path)
        if suffix == ".xlsx":
            return self._extract_xlsx(file_path)

        logger.warning("Формат %s пока не поддерживается для ИИ-анализа: %s", suffix or "без расширения", file_path)
        return ""

    @staticmethod
    def _extract_pdf(file_path: str) -> str:
        try:
            from pypdf import PdfReader

            return "\n".join(page.extract_text() or "" for page in PdfReader(file_path).pages)
        except Exception as error:
            logger.warning("Не удалось извлечь текст из PDF %s: %s", file_path, error)
            return ""

    @staticmethod
    def _extract_xlsx(file_path: str) -> str:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, read_only=True, data_only=True)
            rows = []
            for worksheet in workbook.worksheets:
                rows.append(f"### {worksheet.title} ###")
                rows.extend(
                    " | ".join(str(value) for value in row if value is not None)
                    for row in worksheet.iter_rows(values_only=True)
                    if any(value is not None for value in row)
                )
            workbook.close()
            return "\n".join(rows)
        except Exception as error:
            logger.warning("Не удалось извлечь текст из XLSX %s: %s", file_path, error)
            return ""
